import numpy as np
import pandas as pd
from readdata10000 import readdata10000
from V2QRScalculatornew import V2QRScalculatornew
from V3STEV60calculatornew import V3STEV60calculatornew
from ecgvalueprocessor import ecgvalueprocessor
from V4Rwavecalculatornew0712update import V4Rwavecalculatornew0712update
import os
import openpyxl
import xml.etree.ElementTree as ET

#創造一個資料夾內每個檔案的通用路徑
folderpath = '/Users/huangliangjun/Desktop/醫學五 論文資料/20230823進度/AMIfinal'
allfilelist = os.listdir(folderpath)

wb = openpyxl.Workbook()
ws = wb.active

#填入excel的格式
ws.cell(row = 1, column = 1).value = "filename"
ws.cell(row = 1, column = 2).value = "PatientID"
ws.cell(row = 1, column = 3).value = "AqusitionDate"
ws.cell(row = 1, column = 4).value = "AqusitionTime"
ws.cell(row = 1, column = 5).value = "QTc"
ws.cell(row = 1, column = 6).value = "V2QRS"
ws.cell(row = 1, column = 7).value = "V3STE60"
ws.cell(row = 1, column = 8).value = "RV4"
ws.cell(row = 1, column = 9).value = "Result"
ws.cell(row = 1, column= 10).value = "V2 wavelocation"
ws.cell(row = 1, column= 11).value = "V4 wavelocation"
counter = 2

for file in allfilelist:
        
    path = os.path.join(folderpath,file)
    print(path)
    if ".DS_Store" in path:
        continue
    tree=ET.parse(path)  
    root=tree.getroot()
    
    #撈取測量日期
    for TestDemographics in root.iter('TestDemographics'):#在所有LeadData包含的內容中
        AcquisitionDate = TestDemographics.find('AcquisitionDate').text
    print(AcquisitionDate)
    
    #撈取測量時間
    for TestDemographics in root.iter('TestDemographics'):#在所有LeadData包含的內容中
        AcquisitionTime=TestDemographics.find('AcquisitionTime').text
    print(AcquisitionTime)
    
    for PatientDemographics in root.iter('PatientDemographics'):#在所有LeadData包含的內容中
        PatientID=PatientDemographics.find('PatientID').text
        for RestingECGMeasurements in root.iter('RestingECGMeasurements'):#在所有RestingECGMeasurements包含的內容中
            QTc = int(RestingECGMeasurements.find('QTCorrected').text)#尋找RestingECGMeasurements內的text
            print(QTc)       
                    
        dataextract = readdata10000(path)#這個是class
        V2data = dataextract.V2data10000()
        V3data = dataextract.V3data10000()
        V4data = dataextract.V4data10000()

            #取得halfQRS duration(int)
        for RestingECGMeasurements in root.iter('RestingECGMeasurements'):#打開QRSTimesTypes 的第一層
            DoubleQRSDuration = RestingECGMeasurements.find('QRSDuration').text
            QRSDuration = int(int(DoubleQRSDuration)/2)
        halfQRSDuration = int(QRSDuration/2)
                
        Timeindexlist = []
        for QRSTimesTypes in root.iter('QRSTimesTypes'):#打開QRSTimesTypes 的第一層
            for QRS in root.iter('QRS'):#開QRS(第二層)
                DoubleTime = QRS.find('Time').text
                Time = int(int(DoubleTime)/2)
                Timeindexlist.append(Time)
    
    
        #將V2data還原為原始量值
        newV2data = ecgvalueprocessor(V2data)#這個是class
        newV2arr10000 = newV2data.newecgvalue()
        #取得V2QRS的量值
        V2c = V2QRScalculatornew(newV2arr10000, Timeindexlist, halfQRSDuration)
        V2resultlist = V2c.V2QRS10000()
        V2QRSvalue = V2resultlist[0]
        V2Rwavelocation = V2resultlist[1]
        V2location = V2Rwavelocation*0.002
                
        #將V3data還原為原始量值
        newV3data = ecgvalueprocessor(V3data)#這個是class
        newV3arr10000 = newV3data.newecgvalue()
        #取得V3STE60的量值
        V3c = V3STEV60calculatornew(newV3arr10000, Timeindexlist, halfQRSDuration)
        V3STE60value = V3c.V3QRS10000()   
    

        #將V4data還原為原始量值
        newV4data = ecgvalueprocessor(V4data)#這個是class
        newV4arr10000 = newV4data.newecgvalue()
        #取得RV4量值
        V4c = V4Rwavecalculatornew0712update(newV4arr10000, Timeindexlist, halfQRSDuration)
        RV4resultlist = V4c.V4QRS10000()
        RV4value = RV4resultlist[0]
        Rwavelocation = RV4resultlist[1]
        V4location = Rwavelocation*0.002
                
        #計算公式
        a = 0.052*QTc
        b = 0.151*V2QRSvalue
        c = 0.268*RV4value
        d = 1.062*V3STE60value
        result = a-b-c+d
        print(result)
    
        #把各個資料填到對應的表格裡
        ws.cell(row = counter, column = 1).value = file
        ws.cell(row = counter, column = 2).value = PatientID
        ws.cell(row = counter, column = 3).value = AcquisitionDate
        ws.cell(row = counter, column = 4).value = AcquisitionTime
        ws.cell(row = counter, column = 5).value = QTc
        ws.cell(row = counter, column = 6).value = V2QRSvalue
        ws.cell(row = counter, column = 7).value = V3STE60value
        ws.cell(row = counter, column = 8).value = RV4value
        ws.cell(row = counter, column = 9).value = result
        ws.cell(row = counter, column = 10).value = V2location
        ws.cell(row = counter, column = 11).value = V4location
        wb.save('ecgdataoutput20231020AMI.xlsx')
        counter = counter+1 
        


